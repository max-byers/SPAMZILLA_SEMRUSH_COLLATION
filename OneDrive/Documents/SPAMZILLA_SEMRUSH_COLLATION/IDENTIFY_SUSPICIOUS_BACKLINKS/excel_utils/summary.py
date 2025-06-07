"""
Summary utilities for Excel workbooks.
"""

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def add_summary_to_sheet(workbook, sheet_name, summary_data):
    """
    Add a summary to a worksheet.
    
    Args:
        workbook: The workbook to add the summary to
        sheet_name: The name of the sheet to add the summary to
        summary_data: Dictionary containing summary data
    """
    worksheet = workbook[sheet_name]
    
    # Add title
    title_cell = worksheet.cell(row=1, column=1)
    title_cell.value = "Summary"
    title_cell.font = Font(bold=True, size=14)
    
    # Add summary data
    row = 3
    for key, value in summary_data.items():
        worksheet.cell(row=row, column=1, value=key)
        worksheet.cell(row=row, column=2, value=value)
        row += 1
    
    # Format the summary
    for col in range(1, 3):
        worksheet.column_dimensions[get_column_letter(col)].width = 20
        for row in range(1, row):
            cell = worksheet.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='left', vertical='center')

def add_suspicious_content_summary(workbook, sheet_name, suspicious_data):
    """
    Add a summary of suspicious content to a worksheet.
    
    Args:
        workbook: The workbook to add the summary to
        sheet_name: The name of the sheet to add the summary to
        suspicious_data: Dictionary containing suspicious content data
    """
    worksheet = workbook[sheet_name]
    
    # Add title
    title_cell = worksheet.cell(row=1, column=1)
    title_cell.value = "Suspicious Content Summary"
    title_cell.font = Font(bold=True, size=14)
    
    # Add suspicious content data
    row = 3
    for key, value in suspicious_data.items():
        worksheet.cell(row=row, column=1, value=key)
        worksheet.cell(row=row, column=2, value=value)
        row += 1
    
    # Format the summary
    for col in range(1, 3):
        worksheet.column_dimensions[get_column_letter(col)].width = 20
        for row in range(1, row):
            cell = worksheet.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='left', vertical='center') 