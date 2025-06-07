# core.py
"""
Core Excel workbook creation and manipulation utilities.
"""
import random
import traceback
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

# Import from core utils
try:
    from core.utils import sanitize_sheet_name
except ImportError:
    # Fallback if import fails
    def sanitize_sheet_name(name):
        """Ensure sheet name is valid for Excel"""
        import re
        if len(name) > 31:
            name = name[:31]
        return re.sub(r'[\\/*\[\]:?]', '_', name)

# Import configuration
try:
    from config import FEATURES
except ImportError:
    # Default features if config not available
    FEATURES = {
        'ADD_HYPERLINKS': True,
        'USE_DOMAIN_SPACING': True,
        'HIGHLIGHT_LOST_LINKS': True,
        'ROTATING_DOMAIN_COLORS': True
    }


def create_workbook():
    """
    Create a new Excel workbook.

    Returns:
        An openpyxl Workbook object
    """
    return Workbook()


def create_table(worksheet, data_range, table_name, table_style="TableStyleMedium9"):
    """
    Create an Excel table in the specified worksheet.

    Args:
        worksheet: The openpyxl worksheet
        data_range: Range of cells to include in the table (e.g., "A1:F10")
        table_name: Base name for the table (will be cleaned and made unique)
        table_style: Excel table style to apply

    Returns:
        Boolean indicating success
    """
    try:
        # Clean table name to ensure it's valid
        clean_table_name = ''.join(c for c in table_name if c.isalnum() or c == '_')

        # Ensure the table name is unique and not too long
        if len(clean_table_name) > 250:
            clean_table_name = clean_table_name[:250]

        # Add a unique identifier to avoid duplicate table names
        clean_table_name = f"{clean_table_name}_{random.randint(1000, 9999)}"

        print(f"Creating table: {clean_table_name} with range {data_range}")

        table = Table(displayName=clean_table_name, ref=data_range)
        style = TableStyleInfo(
            name=table_style,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style

        # Add the table to the worksheet
        worksheet.add_table(table)
        return True
    except Exception as e:
        print(f"ERROR creating table: {e}")
        traceback.print_exc()
        return False


def auto_adjust_cells(worksheet, start_row, start_col, end_row, end_col, min_width=10, max_width=60,
                      adjust_height=True):
    """
    Auto-adjust both width and height of cells in the specified range.

    Args:
        worksheet: The openpyxl worksheet
        start_row: Starting row
        start_col: Starting column
        end_row: Ending row
        end_col: Ending column
        min_width: Minimum column width
        max_width: Maximum column width
        adjust_height: Whether to adjust row height
    """
    # Adjust column widths
    for col_idx in range(start_col, end_col + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0

        # Check a sample of values in the column (up to 100 rows)
        for row_idx in range(start_row, min(end_row + 1, start_row + 100)):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            if cell.value:
                cell_text = str(cell.value)
                # Count newlines for height adjustment
                cell_lines = cell_text.count('\n') + 1
                # Get maximum line length for width
                max_line_length = max(len(line) for line in cell_text.split('\n'))
                max_length = max(max_length, max_line_length)

                # Adjust row height if needed and enabled
                if adjust_height and cell_lines > 1:
                    current_height = worksheet.row_dimensions[row_idx].height
                    if current_height is None:
                        current_height = 15  # Default row height
                    needed_height = cell_lines * 15  # Approximately 15 points per line
                    worksheet.row_dimensions[row_idx].height = max(current_height, needed_height)

        # Set column width with padding, bounded by min/max
        adjusted_width = max(min_width, min(max_length + 3, max_width))
        worksheet.column_dimensions[col_letter].width = adjusted_width


def get_alpha_ordered_domains(workbook):
    """
    Get a list of all domains in the workbook, ordered alphabetically.

    Args:
        workbook: The openpyxl workbook

    Returns:
        List of domain names in alphabetical order
    """
    # Extract domain names from sheet names
    domains = set()
    for sheet_name in workbook.sheetnames:
        # Try to extract domain from sheet name (assuming format like domain_Everything)
        parts = sheet_name.split('_')
        if len(parts) >= 2 and parts[1] in ['Everything', 'Quality', 'dodgy']:
            domains.add(parts[0])

    # Sort alphabetically
    return sorted(list(domains))


def find_column_index(worksheet, header_row, column_name):
    """
    Find the index of a column by its header name.

    Args:
        worksheet: The openpyxl worksheet
        header_row: Row where headers are located
        column_name: Name of the column to find

    Returns:
        Column index (1-based) or None if not found
    """
    for col_idx in range(1, worksheet.max_column + 1):
        cell_value = worksheet.cell(row=header_row, column=col_idx).value
        if cell_value == column_name:
            return col_idx
    return None