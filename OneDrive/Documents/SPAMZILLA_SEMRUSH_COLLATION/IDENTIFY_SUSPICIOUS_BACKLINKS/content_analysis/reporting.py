# content_analysis/reporting.py
"""
Functions for generating reports and Excel sheets for suspicious content analysis.
"""
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# Import from excel_utils package
try:
    from excel_utils import add_df_to_worksheet
    from excel_utils.navigating import add_navigation_links
    from excel_utils.core import auto_adjust_cells
except ImportError:
    # Fallback paths during development
    try:
        from excel_utils.dataframe import add_df_to_worksheet
        from excel_utils.navigating import add_navigation_links
        from excel_utils.core import auto_adjust_cells
    except ImportError:
        print("Error importing excel_utils functions. Check package structure.")


        # Define stub functions to prevent crashes
        def add_df_to_worksheet(*args, **kwargs):
            pass


        def add_navigation_links(*args, **kwargs):
            pass


        def auto_adjust_cells(*args, **kwargs):
            pass

# Import from core utils
from core.utils import sanitize_sheet_name

# Import from our own modules
from content_analysis.utils import get_alpha_ordered_domains
from content_analysis.detection import analyze_content


def create_valid_sheet_name(domain_name, suffix):
    """
    Create a valid Excel sheet name (max 31 chars) from a domain name and suffix.
    
    Args:
        domain_name: The domain name
        suffix: The suffix to add (e.g., '_dodgy', '_Everything')
    
    Returns:
        A valid Excel sheet name
    """
    # Remove common TLDs to save space
    domain = domain_name.replace('.com', '').replace('.net', '').replace('.org', '')
    
    # Create base name
    base_name = f"{domain}{suffix}"
    
    # If still too long, truncate
    if len(base_name) > 31:
        # Keep first 28 chars and add '...'
        base_name = base_name[:28] + '...'
    
    return base_name


def add_dodgy_domain_sheet(workbook, domain_name, dodgy_data, domain_everything_df=None, sheet_number=None):
    """
    Add a domain-specific sheet for dodgy keywords with improved formatting and navigation.
    Filters out backlinks with external links > 5000.

    Args:
        workbook: The openpyxl workbook
        domain_name: The domain name for this sheet
        dodgy_data: DataFrame with dodgy keyword data
        domain_everything_df: Optional original DataFrame to get External Links values
        sheet_number: The sheet number for this domain

    Returns:
        The created worksheet
    """
    # Create valid sheet name
    sheet_name = create_valid_sheet_name(domain_name, "_dodgy")

    # Create the sheet
    try:
        print(f"\nDebug - {domain_name} dodgy data columns before processing:")
        print(dodgy_data.columns.tolist())

        # Filter out backlinks with external links > 5000
        if 'External Links' in dodgy_data.columns:
            original_count = len(dodgy_data)
            dodgy_data = dodgy_data[dodgy_data['External Links'].astype(float) <= 5000]
            filtered_count = len(dodgy_data)
            print(f"\nDebug - {domain_name}: Filtered out {original_count - filtered_count} backlinks with > 5000 external links")
            print(f"Debug - {domain_name}: Remaining suspicious backlinks: {filtered_count}")

        dodgy_sheet = workbook.create_sheet(title=sheet_name)

        # Add navigation links - only one link to Summary in A1
        add_navigation_links(dodgy_sheet, "Summary", 1, 1)

        # Get all domain names in the workbook (for navigation)
        domain_names = get_alpha_ordered_domains(workbook)

        # Find the current domain position in alphabetical order
        current_idx = -1
        try:
            current_idx = domain_names.index(domain_name)
        except ValueError:
            pass  # Domain not found in list

        # If we found the domain, add navigation between dodgy sheets
        if current_idx >= 0 and len(domain_names) > 1:
            # Determine previous domain (wrap around if at start)
            prev_idx = current_idx - 1 if current_idx > 0 else len(domain_names) - 1
            prev_domain = domain_names[prev_idx]
            prev_sheet_name = f"{prev_domain}-backlinks_dodgy"

            # Determine next domain (wrap around if at end)
            next_idx = (current_idx + 1) % len(domain_names)
            next_domain = domain_names[next_idx]
            next_sheet_name = f"{next_domain}-backlinks_dodgy"

            # Add Previous link
            if prev_sheet_name in workbook.sheetnames:
                prev_cell = dodgy_sheet.cell(row=1, column=3, value=f"« Previous Dodgy ({prev_domain})")
                prev_cell.hyperlink = f"#{prev_sheet_name}!A1"
                prev_cell.font = Font(color="0000FF", underline="single")

            # Add Next link
            if next_sheet_name in workbook.sheetnames:
                next_cell = dodgy_sheet.cell(row=1, column=5, value=f"Next Dodgy ({next_domain}) »")
                next_cell.hyperlink = f"#{next_sheet_name}!A1"
                next_cell.font = Font(color="0000FF", underline="single")

            # Add Content Analysis link
            add_navigation_links(dodgy_sheet, "Content Analysis", 1, 7)
        else:
            # If no navigation links, add Content Analysis link in column 3 instead
            add_navigation_links(dodgy_sheet, "Content Analysis", 1, 3)

        # Add content if we have dodgy data
        if not dodgy_data.empty:
            print(f"\nDebug - {domain_name} dodgy data columns after navigation:")
            print(dodgy_data.columns.tolist())

            # Add External Links column from the everything data if available
            if domain_everything_df is not None and 'External links' in domain_everything_df.columns:
                # Create a mapping from Source URL to External links count
                url_to_external_links = {}
                source_url_col = 'Source url' if 'Source url' in domain_everything_df.columns else 'Source URL'

                if source_url_col in domain_everything_df.columns:
                    for _, row in domain_everything_df.iterrows():
                        url = str(row[source_url_col])
                        if 'External links' in row:
                            url_to_external_links[url] = row['External links']

                    print(f"\nDebug - {domain_name} url_to_external_links mapping created with {len(url_to_external_links)} entries")

                # Add External Links column to dodgy_data if it doesn't exist
                if 'External Links' not in dodgy_data.columns:
                    external_links = []
                    source_url_col = 'Source URL' if 'Source URL' in dodgy_data.columns else 'Source url'

                    if source_url_col in dodgy_data.columns:
                        for _, row in dodgy_data.iterrows():
                            url = str(row[source_url_col])
                            external_links.append(url_to_external_links.get(url, ""))

                        print(f"\nDebug - {domain_name} external_links list created with {len(external_links)} entries")

                        # Insert after Flagged Word(s) column if it exists
                        if 'Flagged Word(s)' in dodgy_data.columns:
                            flagged_idx = dodgy_data.columns.get_loc('Flagged Word(s)')
                            dodgy_data.insert(flagged_idx + 1, 'External Links', external_links)
                        else:
                            # Otherwise add at the end
                            dodgy_data['External Links'] = external_links

            print(f"\nDebug - {domain_name} dodgy data columns before reordering:")
            print(dodgy_data.columns.tolist())

            # Ensure columns are in the correct order
            column_order = [
                'Referring Domain',
                'Flagged Word(s)',
                'Categories',
                'Source URL',
                'Anchor',
                'Title',
                'External Links',
                'Location'
            ]
            
            # Add any missing columns with empty values
            for col in column_order:
                if col not in dodgy_data.columns:
                    dodgy_data[col] = ""

            print(f"\nDebug - {domain_name} dodgy data columns after adding missing:")
            print(dodgy_data.columns.tolist())

            # Reorder columns
            dodgy_data = dodgy_data[column_order]

            print(f"\nDebug - {domain_name} dodgy data columns after reordering:")
            print(dodgy_data.columns.tolist())

            # Add the dataframe
            start_row = 3
            add_df_to_worksheet(
                dodgy_sheet,
                dodgy_data,
                start_row=start_row,
                start_col=1,
                table_name=f"{domain_name}_dodgy_kw",
                title=f"Suspicious Keywords for {domain_name}"
            )

            # Highlight suspicious content in its original location
            highlight_suspicious_content(
                dodgy_sheet,
                start_row,
                dodgy_data,
                'Flagged Word(s)',
                'Categories',
                'Location'
            )

            # Add hyperlinks to the Everything sheet from the row numbers
            if 'Row Number' in dodgy_data.columns:
                row_idx = dodgy_data.columns.get_loc('Row Number') + 1
                everything_sheet_name = f"{domain_name}-backlinks_Everything"

                for data_row in range(4, 4 + len(dodgy_data)):
                    cell = dodgy_sheet.cell(row=data_row, column=row_idx)
                    row_num = cell.value

                    if row_num and everything_sheet_name in workbook.sheetnames:
                        # Excel rows are 1-based, and we need to account for headers
                        target_row = row_num + 1  # +1 for header row
                        cell.hyperlink = f"#{everything_sheet_name}!A{target_row}"
                        cell.font = Font(color="0000FF", underline="single")
                        cell.value = f"Row {row_num}"

        return dodgy_sheet
    except Exception as e:
        print(f"Error creating dodgy sheet for {domain_name}: {e}")
        import traceback
        traceback.print_exc()
        return None


def highlight_suspicious_content(worksheet, start_row, df, flagged_words_col, categories_col, locations_col):
    """
    Highlight suspicious content in all relevant columns based on flagged words.

    Args:
        worksheet: The openpyxl worksheet
        start_row: Starting row of the data (header row)
        df: DataFrame with suspicious content data
        flagged_words_col: Column name containing flagged words
        categories_col: Column name containing categories
        locations_col: Column name containing locations (which columns have flagged content)

    Returns:
        Number of cells highlighted
    """
    highlighted_count = 0

    # Check if required columns exist
    if flagged_words_col not in df.columns or locations_col not in df.columns:
        return 0

    # Map column names to their indices in the worksheet
    col_indices = {}
    for col_idx in range(1, worksheet.max_column + 1):
        header_value = worksheet.cell(row=start_row, column=col_idx).value
        if header_value:
            col_indices[header_value] = col_idx

    # Process each row in the dataframe
    for row_idx, row in enumerate(df.itertuples(), 1):
        # Skip header row
        if row_idx == 1:
            continue

        # Extract flagged words and locations
        flagged_words_attr = flagged_words_col.replace(' ', '_')
        locations_attr = locations_col.replace(' ', '_')

        if hasattr(row, flagged_words_attr) and hasattr(row, locations_attr):
            flagged_words = getattr(row, flagged_words_attr)
            locations = getattr(row, locations_attr)

            if pd.isna(flagged_words) or pd.isna(locations):
                continue

            # Split into lists
            word_list = [w.strip().lower() for w in str(flagged_words).split(',')]
            location_list = [l.strip() for l in str(locations).split(',')]

            # Get the worksheet row number (header + data row offset)
            ws_row = start_row + row_idx

            # For each mentioned location, check and highlight the cell
            for location in location_list:
                if location in col_indices:
                    col_idx = col_indices[location]
                    cell = worksheet.cell(row=ws_row, column=col_idx)

                    if cell.value:
                        cell_text = str(cell.value).lower()

                        # Check if any flagged word is in this cell
                        if any(word in cell_text for word in word_list):
                            cell.font = Font(color="FF0000", bold=True)
                            highlighted_count += 1

    return highlighted_count


def add_content_analysis_sheet(workbook, domain_data_dict, domain_to_sheet_number):
    """
    Add a content analysis sheet to the workbook based on all domain data.

    Args:
        workbook: The openpyxl workbook
        domain_data_dict: Dictionary mapping domain names to their dataframes
        domain_to_sheet_number: Dictionary mapping domains to their sheet numbers

    Returns:
        A tuple containing (content_analysis_sheet, domain_dodgy_count_dict, domain_dodgy_domains_dict)
    """
    # Create a new sheet for content analysis
    content_sheet = workbook.create_sheet(title="Content Analysis")

    # Add a back to summary link
    add_navigation_links(workbook, content_sheet, "Summary", (1, 1))

    # Add a title
    content_sheet.cell(row=1, column=3, value="BACKLINK CONTENT ANALYSIS - SUSPICIOUS KEYWORD DETECTION")
    content_sheet.cell(row=1, column=3).font = Font(bold=True, size=14)
    content_sheet.merge_cells('C1:H1')
    content_sheet.cell(row=1, column=3).alignment = Alignment(horizontal='center')

    # Dictionary to track dodgy keyword counts by domain
    domain_dodgy_count_dict = {}
    domain_dodgy_domains_dict = {}

    # Process each domain
    all_results = []

    # Get alphabetically sorted domain list
    domain_names = sorted(domain_data_dict.keys(), key=str.lower)

    for domain in domain_names:
        df = domain_data_dict[domain]
        sheet_number = domain_to_sheet_number.get(domain)

        # Skip empty dataframes or domains without sheet numbers
        if df is None or df.empty or sheet_number is None:
            continue

        print(f"Analyzing content for domain: {domain}")

        # Analyze content
        domain_results = analyze_content(df, domain_name=domain)

        # If dodgy keywords found, create a domain-specific dodgy sheet
        if not domain_results.empty:
            add_dodgy_domain_sheet(workbook, domain, domain_results, df, sheet_number)

            # Count the number of dodgy backlinks
            dodgy_count = len(domain_results)
            domain_dodgy_count_dict[domain] = dodgy_count

            # Count the number of unique referring domains with dodgy backlinks
            unique_dodgy_domains = domain_results['Referring Domain'].nunique()
            domain_dodgy_domains_dict[domain] = unique_dodgy_domains

            # Add to overall results
            all_results.append(domain_results)
        else:
            domain_dodgy_count_dict[domain] = 0
            domain_dodgy_domains_dict[domain] = 0

    # Combine all results
    if all_results:
        combined_results = pd.concat(all_results, ignore_index=True)

        # Sort by domain and amount (descending)
        combined_results = combined_results.sort_values(['Domain', 'Amount'], ascending=[True, False])

        # Add the results to the sheet
        add_df_to_worksheet(
            content_sheet,
            combined_results,
            start_row=3,start_col=1,
            table_name="suspicious_content",
            title="Detected Suspicious Content"
        )

        # Modify cell formatting to make flagged words red and bold
        flagged_col_idx = combined_results.columns.get_loc('Flagged Word(s)') + 1
        for row_idx in range(4, 4 + len(combined_results)):
            # Make the flagged words cell red and bold
            cell = content_sheet.cell(row=row_idx, column=flagged_col_idx)
            cell.font = Font(color="FF0000", bold=True)

            # Apply light red background to rows with Amount > 1
            amount_col_idx = combined_results.columns.get_loc('Amount') + 1
            amount_cell = content_sheet.cell(row=row_idx, column=amount_col_idx)

            try:
                if amount_cell.value and int(str(amount_cell.value)) > 1:
                    # Apply light red fill to the entire row
                    for col_idx in range(1, len(combined_results.columns) + 1):
                        content_sheet.cell(row=row_idx, column=col_idx).fill = PatternFill(
                            start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            except (ValueError, TypeError):
                # Skip if amount isn't a valid number
                pass

        # Add hyperlinks from domain names to their dodgy sheets
        domain_col_idx = combined_results.columns.get_loc('Domain') + 1
        for row_idx in range(4, 4 + len(combined_results)):
            domain_cell = content_sheet.cell(row=row_idx, column=domain_col_idx)
            domain_name = domain_cell.value

            # Create the hyperlink to the dodgy sheet
            if domain_name:
                dodgy_sheet_name = f"{domain_name}-backlinks_dodgy"

                if dodgy_sheet_name in workbook.sheetnames:
                    domain_cell.hyperlink = f"#{dodgy_sheet_name}!A1"
                    domain_cell.font = Font(color="0000FF", underline="single")

        # Add hyperlinks to source URLs
        source_url_col_idx = combined_results.columns.get_loc('Source URL') + 1
        for row_idx in range(4, 4 + len(combined_results)):
            url_cell = content_sheet.cell(row=row_idx, column=source_url_col_idx)
            url = url_cell.value

            if url and isinstance(url, str) and url.startswith('http'):
                url_cell.hyperlink = url
                url_cell.font = Font(color="0000FF", underline="single")

        # Auto-adjust for better fit
        auto_adjust_cells(
            content_sheet,
            3,  # start_row
            1,  # start_col
            3 + len(combined_results),  # end_row
            len(combined_results.columns),  # end_col
            min_width=15,
            max_width=60,
            adjust_height=True
        )

    else:
        # Add a message if no suspicious content was found
        content_sheet.cell(row=3, column=1, value="No suspicious content detected in any domain")
        content_sheet.cell(row=3, column=1).font = Font(italic=True)

    return content_sheet, domain_dodgy_count_dict, domain_dodgy_domains_dict


def create_suspicious_summary_report(workbook, domain_data_dict):
    """
    Create a comprehensive summary report of suspicious content across all domains.

    Args:
        workbook: The openpyxl workbook
        domain_data_dict: Dictionary mapping domain names to their dataframes

    Returns:
        DataFrame with suspicious content summary
    """
    # Get summary data for each domain
    summary_data = []

    # Get alphabetically sorted domain list
    domain_names = sorted(domain_data_dict.keys(), key=str.lower)

    for domain in domain_names:
        df = domain_data_dict[domain]

        # Skip empty dataframes
        if df is None or df.empty:
            continue

        # Get suspicious content
        suspicious_df = analyze_content(df, domain_name=domain)

        # Calculate metrics
        total_backlinks = len(df)
        suspicious_backlinks = len(suspicious_df)

        # Get domain counts
        total_domains = df['Referring Domain'].nunique() if 'Referring Domain' in df.columns else 0
        suspicious_domains = suspicious_df['Referring Domain'].nunique() if not suspicious_df.empty else 0

        # Get category breakdown
        category_counts = {}
        if not suspicious_df.empty and 'Categories' in suspicious_df.columns:
            all_categories = []
            for categories in suspicious_df['Categories']:
                if isinstance(categories, str):
                    all_categories.extend([cat.strip() for cat in categories.split(',')])

            for cat in all_categories:
                category_counts[cat] = category_counts.get(cat, 0) + 1

        # Add to summary data
        summary_data.append({
            'Domain': domain,
            'Total Backlinks': total_backlinks,
            'Suspicious Backlinks': suspicious_backlinks,
            'Suspicious %': round(suspicious_backlinks / max(1, total_backlinks) * 100, 1),
            'Total Referring Domains': total_domains,
            'Suspicious Domains': suspicious_domains,
            'Domain Coverage %': round(suspicious_domains / max(1, total_domains) * 100, 1),
            'Categories': ', '.join(f"{cat} ({count})" for cat, count in category_counts.items())
        })

    # Create DataFrame and sort alphabetically by domain
    summary_df = pd.DataFrame(summary_data)
    if not summary_df.empty:
        summary_df = summary_df.sort_values('Domain', ascending=True)

    return summary_df
