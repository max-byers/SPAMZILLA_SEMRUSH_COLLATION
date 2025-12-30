# excel_utils/formatting.py
"""
Excel cell and worksheet formatting utilities.
"""
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import Rule

# Import from core constants if available
try:
    from core.constants import DOMAIN_COLORS
except ImportError:
    # Default colors if import fails
    DOMAIN_COLORS = [
        "0000FF",  # Blue
        "FF0000",  # Red
        "008000",  # Green
        "800080",  # Purple
        "FFA500",  # Orange
        "008080",  # Teal
    ]


def add_title_only(worksheet, title, start_row, start_col, num_columns=5):
    """
    Add just a title to the worksheet when there's no data to display.

    Args:
        worksheet: The openpyxl worksheet
        title: Title text
        start_row: Starting row for the title
        start_col: Starting column for the title
        num_columns: Number of columns to merge for the title

    Returns:
        Next row after the title (with gap)
    """
    worksheet.cell(row=start_row, column=start_col, value=title)
    worksheet.cell(row=start_row, column=start_col).font = Font(bold=True, size=14)
    worksheet.cell(row=start_row, column=start_col).alignment = Alignment(horizontal='center')

    # Merge cells for the title across columns
    end_col = start_col + num_columns - 1
    worksheet.merge_cells(start_row=start_row, start_column=start_col,
                          end_row=start_row, end_column=end_col)

    return start_row + 2  # Return next row after title with a small gap


def add_black_divider_column(worksheet, col_idx, start_row, end_row, width=2):
    """
    Add a black divider column to visually separate sections.

    Args:
        worksheet: The openpyxl worksheet
        col_idx: Column index (1-based)
        start_row: Start row (1-based)
        end_row: End row (1-based)
        width: Width of the divider column in characters (default 2)
    """
    # Create black fill
    black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

    # Set column width to be narrow
    col_letter = get_column_letter(col_idx)
    worksheet.column_dimensions[col_letter].width = width

    # Fill cells with black
    for row in range(start_row, end_row + 1):
        cell = worksheet.cell(row=row, column=col_idx)
        cell.fill = black_fill
        # Clear any existing value
        cell.value = None


def apply_cell_highlighting(worksheet, row, col, color="FFFF00", text_color=None):
    """
    Apply background highlighting to a cell.

    Args:
        worksheet: The openpyxl worksheet
        row: Row index (1-based)
        col: Column index (1-based)
        color: Hex color code for background
        text_color: Hex color code for text (optional)
    """
    cell = worksheet.cell(row=row, column=col)
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    if text_color:
        cell.font = Font(color=text_color)


def apply_row_highlighting(worksheet, row, start_col, end_col, color="FFFF00", text_color=None):
    """
    Apply background highlighting to an entire row.

    Args:
        worksheet: The openpyxl worksheet
        row: Row index (1-based)
        start_col: Starting column index (1-based)
        end_col: Ending column index (inclusive, 1-based)
        color: Hex color code for background
        text_color: Hex color code for text (optional)
    """
    for col in range(start_col, end_col + 1):
        apply_cell_highlighting(worksheet, row, col, color, text_color)


def apply_alternating_row_colors(worksheet, start_row, end_row, start_col, end_col,
                                 color1="FFFFFF", color2="F0F0F0"):
    """
    Apply alternating row colors to a range of cells.

    Args:
        worksheet: The openpyxl worksheet
        start_row: Starting row index (1-based)
        end_row: Ending row index (inclusive, 1-based)
        start_col: Starting column index (1-based)
        end_col: Ending column index (inclusive, 1-based)
        color1: Hex color code for odd rows
        color2: Hex color code for even rows
    """
    for row in range(start_row, end_row + 1):
        color = color1 if (row - start_row) % 2 == 0 else color2
        for col in range(start_col, end_col + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")


def apply_header_formatting(worksheet, row, start_col, end_col,
                            bg_color="D0D0D0", text_color="000000", bold=True):
    """
    Apply header formatting to a row.

    Args:
        worksheet: The openpyxl worksheet
        row: Row index (1-based)
        start_col: Starting column index (1-based)
        end_col: Ending column index (inclusive, 1-based)
        bg_color: Hex color code for background
        text_color: Hex color code for text
        bold: Whether text should be bold
    """
    for col in range(start_col, end_col + 1):
        cell = worksheet.cell(row=row, column=col)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        cell.font = Font(bold=bold, color=text_color)
        cell.alignment = worksheet.cell(row=row, column=col).alignment


def apply_domain_color_rotation(worksheet, df, domain_col_idx, start_row, start_col):
    """
    Apply rotating colors to domains in the worksheet.

    Args:
        worksheet: The openpyxl worksheet
        df: The pandas DataFrame containing the data
        domain_col_idx: The column index of the Referring Domain column
        start_row: Starting row for the data (after headers)
        start_col: Starting column for the data

    Returns:
        Dictionary mapping domains to their assigned colors
    """
    # Track domains we've seen and their assigned colors
    domain_color_map = {}
    current_color_idx = 0

    # Apply coloring to each domain
    for row_idx, domain in enumerate(df['Referring Domain']):
        if domain not in domain_color_map:
            # Assign the next color in rotation
            domain_color_map[domain] = DOMAIN_COLORS[current_color_idx % len(DOMAIN_COLORS)]
            current_color_idx += 1

        # Apply the color to the domain cell
        cell = worksheet.cell(row=start_row + row_idx, column=start_col + domain_col_idx)
        cell.font = Font(color=domain_color_map[domain])

    return domain_color_map


def highlight_flagged_words(worksheet, row, col, text, keywords, color="FF0000", make_bold=True):
    """
    Highlight flagged keywords in text by making them red and bold.
    Uses rich text formatting in the cell.

    Args:
        worksheet: The openpyxl worksheet
        row: Row index (1-based)
        col: Column index (1-based)
        text: The text content
        keywords: List of keywords to highlight
        color: Hex color code for the highlighted words
        make_bold: Whether to also make the text bold

    Note:
        This is a placeholder - full rich text formatting requires
        more complex implementation with openpyxl's rich text features.
        In practice, this would involve creating Rich Text runs.
    """
    # This is a basic implementation that just sets the whole cell
    # A full implementation would use openpyxl's rich text features
    cell = worksheet.cell(row=row, column=col, value=text)

    # If any keywords are found in the text, apply highlighting to the whole cell
    if any(keyword.lower() in text.lower() for keyword in keywords):
        font_props = {}
        if make_bold:
            font_props['bold'] = True
        if color:
            font_props['color'] = color

        cell.font = Font(**font_props)


def add_last_seen_conditional_formatting(worksheet, col_letter, start_row, end_row):
    """
    Add conditional formatting to highlight the Last seen column based on date.

    Args:
        worksheet: The openpyxl worksheet
        col_letter: Column letter (e.g., 'G')
        start_row: Starting row (after header)
        end_row: Ending row
    """
    range_str = f"{col_letter}{start_row}:{col_letter}{end_row}"

    try:
        # Green (within day)
        worksheet.conditional_formatting.add(
            range_str,
            Rule(type='timePeriod', timePeriod='yesterday', dxf=None, priority=1)
        )

        # Yellow (within week)
        worksheet.conditional_formatting.add(
            range_str,
            Rule(type='timePeriod', timePeriod='lastWeek', dxf=None, priority=2)
        )

        # Orange (within month)
        worksheet.conditional_formatting.add(
            range_str,
            Rule(type='timePeriod', timePeriod='lastMonth', dxf=None, priority=3)
        )

        # Red (earlier than a month)
        worksheet.conditional_formatting.add(
            range_str,
            Rule(type='timePeriod', timePeriod='lastMonth', operator='notEqual', dxf=None, priority=4)
        )
    except Exception as e:
        print(f"Warning: Could not add time-based conditional formatting: {e}")


def apply_total_row_formatting(worksheet, row, start_col, end_col):
    """
    Format a total row with borders and bold text.

    Args:
        worksheet: The openpyxl worksheet
        row: Row index for total row
        start_col: Starting column
        end_col: Ending column
    """
    # Add medium border above total row
    border_top = Border(top=Side(style='medium'))

    for col in range(start_col, end_col + 1):
        cell = worksheet.cell(row=row, column=col)
        cell.border = border_top
        cell.font = Font(bold=True)