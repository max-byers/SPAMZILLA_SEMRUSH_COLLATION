# excel_utils/dataframe.py
"""
Utilities for converting pandas DataFrames to Excel worksheets with formatting.
"""
import traceback
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.comments import Comment

# Import from our excel_utils package
try:
    from excel_utils.core import create_table, auto_adjust_cells
    from excel_utils.formatting import add_title_only, apply_domain_color_rotation
    from excel_utils.formatting import add_last_seen_conditional_formatting
except ImportError:
    # Fallback to direct imports during development
    from excel_core import create_table, auto_adjust_cells
    from excel_formatting import add_title_only, apply_domain_color_rotation
    from excel_formatting import add_last_seen_conditional_formatting

# Import configuration
try:
    from config import FEATURES
except ImportError:
    # Default features if config not available
    FEATURES = {
        'ADD_HYPERLINKS': True,
        'USE_DOMAIN_SPACING': True,
        'HIGHLIGHT_LOST_LINKS': True,
        'ROTATING_DOMAIN_COLORS': True
    }


def add_df_to_worksheet(worksheet, df, start_row=1, start_col=1, create_table_flag=True, table_name=None, title=None,
                        highlight_keywords=None):
    """
    Add a DataFrame to a worksheet with comprehensive formatting.

    Args:
        worksheet: The openpyxl worksheet
        df: pandas DataFrame to add
        start_row: Starting row index
        start_col: Starting column index
        create_table_flag: Whether to create an Excel table
        table_name: Name for the table
        title: Title to display above the data
        highlight_keywords: Dictionary of {column: [keywords]} to highlight

    Returns:
        Next available row index after the added data
    """
    try:
        print(f"Adding DataFrame: rows={len(df)} cols={len(df.columns) if not df.empty else 0}")

        # Handle empty DataFrame case
        if df.empty:
            print(f"DataFrame is empty. Adding title only: {title}")
            if title:
                return add_title_only(worksheet, f"{title} (No data in this category)", start_row, start_col)
            return start_row

        # If "Nofollow" column is present, remove it after filtering
        if 'Nofollow' in df.columns:
            df = df.drop(columns=['Nofollow'])

        # Add title if provided
        current_row = start_row
        if title:
            worksheet.cell(row=current_row, column=start_col, value=title)
            worksheet.cell(row=current_row, column=start_col).font = Font(bold=True, size=14)
            worksheet.cell(row=current_row, column=start_col).alignment = Alignment(horizontal='center')

            # Merge cells for the title across all columns
            if len(df.columns) > 1:
                end_col = start_col + len(df.columns) - 1
                worksheet.merge_cells(start_row=current_row, start_column=start_col,
                                      end_row=current_row, end_column=end_col)
            current_row += 1

        # Find the domain column index for later use
        domain_col_idx = -1
        if 'Referring Domain' in df.columns:
            domain_col_idx = df.columns.get_loc('Referring Domain')

        # Extract columns that need keyword highlighting
        highlight_columns = {}
        if highlight_keywords:
            for col_name, keywords in highlight_keywords.items():
                if col_name in df.columns:
                    col_idx = df.columns.get_loc(col_name)
                    highlight_columns[col_idx] = keywords

        # Convert DataFrame to rows and write to worksheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
            for c_idx, value in enumerate(row):
                cell = worksheet.cell(row=current_row + r_idx, column=start_col + c_idx, value=value)

                # Make header row bold
                if r_idx == 0:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                    continue

                # Apply formatting for data rows
                # Bold the Referring Domain column (data cells)
                if c_idx == domain_col_idx:
                    cell.font = Font(bold=True)

                # Add hyperlinks for URL columns
                if df.columns[c_idx] in ['Source url', 'Source URL'] and value and isinstance(value, str):
                    if value.startswith('http') and FEATURES.get('ADD_HYPERLINKS', True):
                        cell.hyperlink = value
                        cell.font = Font(color="0000FF", underline="single")

                # Highlight keywords in specific columns
                if c_idx in highlight_columns and value and isinstance(value, str):
                    keywords = highlight_columns[c_idx]
                    for keyword in keywords:
                        if keyword.lower() in value.lower():
                            # Highlight the keyword
                            current_font = cell.font
                            cell.font = Font(
                                name=current_font.name if current_font.name else 'Calibri',
                                size=current_font.size,
                                bold=True,
                                color="FF0000"  # Red color for flagged words
                            )
                            break

                # Add domain spacing (alternating row background when domain changes)
                if (r_idx > 1 and domain_col_idx >= 0 and c_idx == domain_col_idx
                        and FEATURES.get('USE_DOMAIN_SPACING', True)):
                    # Get current and previous domain
                    current_domain = value
                    prev_row_idx = r_idx - 1
                    prev_domain = worksheet.cell(
                        row=current_row + prev_row_idx,
                        column=start_col + c_idx
                    ).value

                    # If domain changed, add background to the new domain row
                    if current_domain != prev_domain:
                        for col_idx in range(len(df.columns)):
                            worksheet.cell(row=current_row + r_idx, column=start_col + col_idx).fill = PatternFill(
                                start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')

        # Apply rotating colors to domains if feature is enabled
        if domain_col_idx >= 0 and FEATURES.get('ROTATING_DOMAIN_COLORS', True):
            apply_domain_color_rotation(worksheet, df, domain_col_idx, current_row + 1, start_col)

        # Create table if requested
        if create_table_flag:
            # Calculate table range - headers + data rows
            end_row = current_row + len(df)
            end_col = start_col + len(df.columns) - 1
            table_range = f"{get_column_letter(start_col)}{current_row}:{get_column_letter(end_col)}{end_row}"

            # Create the table
            table_created = create_table(worksheet, table_range, table_name if table_name else "Table")

            # Add conditional formatting for Last seen column
            if 'Last seen' in df.columns and table_created:
                try:
                    last_seen_idx = df.columns.get_loc('Last seen') + start_col
                    col_letter = get_column_letter(last_seen_idx)
                    # Add time-based formatting to Last seen column
                    add_last_seen_conditional_formatting(
                        worksheet,
                        col_letter,
                        current_row + 1,  # Start after header
                        end_row
                    )
                except Exception as e:
                    print(f"Error adding Last seen conditional formatting: {e}")

            # Add highlighting for lost links
            if 'Lost link' in df.columns and table_created and FEATURES.get('HIGHLIGHT_LOST_LINKS', True):
                try:
                    lost_idx = df.columns.get_loc('Lost link') + start_col
                    for row_idx in range(current_row + 1, end_row + 1):
                        cell = worksheet.cell(row=row_idx, column=lost_idx)
                        # Check if the cell has a value indicating a lost link
                        if cell.value and str(cell.value).lower() not in ('none', 'null', '', 'na', 'false'):
                            # Highlight entire row
                            for col_idx in range(start_col, start_col + len(df.columns)):
                                worksheet.cell(row=row_idx, column=col_idx).fill = PatternFill(
                                    start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')

                            # Optionally add a comment with the reason for lost status
                            try:
                                comment_text = f"Lost link: {cell.value}"
                                comment = Comment(comment_text, "Backlink Tool")
                                worksheet.cell(row=row_idx, column=start_col).comment = comment
                            except Exception as e:
                                print(f"Warning: Could not add comment: {e}")
                except Exception as e:
                    print(f"Error highlighting lost links: {e}")

        # Auto-adjust cells for better fit
        auto_adjust_cells(
            worksheet,
            current_row,
            start_col,
            current_row + len(df),
            start_col + len(df.columns) - 1
        )

        # Return next available row
        return current_row + len(df) + 1

    except Exception as e:
        print(f"Error in add_df_to_worksheet: {e}")
        traceback.print_exc()
        if title:
            return add_title_only(worksheet, f"{title} (Error displaying data)", start_row, start_col)
        return start_row + 2


def highlight_suspicious_content(worksheet, start_row, df, flagged_words_col, categories_col, locations_col):
    """
    Highlight suspicious content in all relevant columns based on flagged words.

    Args:
        worksheet: The openpyxl worksheet
        start_row: Starting row of the data (header row)
        df: DataFrame with suspicious content data
        flagged_words_col: Column name containing flagged words
        categories_col: Column name containing categories
        locations_col: Column name containing locations (which columns have flagged content)

    Returns:
        Number of cells highlighted
    """
    highlighted_count = 0

    # Check if required columns exist
    if flagged_words_col not in df.columns or locations_col not in df.columns:
        return 0

    # Process each row
    for r_idx, row in enumerate(df.itertuples(), 1):
        # Skip header row
        if r_idx == 1:
            continue

        # Get flagged words and their locations
        row_idx = getattr(row, 'Index')
        if hasattr(row, flagged_words_col.replace(' ', '_')):
            flagged_words = getattr(row, flagged_words_col.replace(' ', '_'))
            locations = getattr(row, locations_col.replace(' ', '_'))

            if pd.isna(flagged_words) or pd.isna(locations):
                continue

            # Split into lists
            word_list = [w.strip() for w in str(flagged_words).split(',')]
            location_list = [l.strip() for l in str(locations).split(',')]

            # For each relevant column, apply highlighting
            actual_row = start_row + r_idx
            for col_name in location_list:
                col_idx = None

                # Find column index
                for c_idx, name in enumerate(df.columns):
                    if name.lower() == col_name.lower():
                        col_idx = c_idx + 1  # 1-based indexing for worksheet
                        break

                if col_idx is not None:
                    cell = worksheet.cell(row=actual_row, column=col_idx)

                    # Apply red, bold formatting to the cell
                    cell.font = Font(color="FF0000", bold=True)
                    highlighted_count += 1

    return highlighted_count


def add_external_links_to_dodgy_sheet(worksheet, df, start_row, original_df):
    """
    Add the External Links column to a dodgy sheet by looking up values in the original data.

    Args:
        worksheet: The openpyxl worksheet (dodgy sheet)
        df: DataFrame for the dodgy sheet
        start_row: Starting row of the data (after header)
        original_df: Original DataFrame containing the External Links column

    Returns:
        Boolean indicating success
    """
    # Check if the required columns exist
    if 'Source URL' not in df.columns or 'External links' not in original_df.columns:
        return False

    # Find or create External Links column in the dodgy sheet
    ext_links_col = None
    for col_idx in range(1, worksheet.max_column + 1):
        if worksheet.cell(row=start_row, column=col_idx).value == 'External Links':
            ext_links_col = col_idx
            break

    # Create the column if it doesn't exist
    if ext_links_col is None:
        ext_links_col = worksheet.max_column + 1
        worksheet.cell(row=start_row, column=ext_links_col, value='External Links').font = Font(bold=True)

    # Create a lookup dictionary from original_df
    source_url_to_ext_links = dict(zip(
        original_df['Source url'].astype(str),
        original_df['External links']
    ))

    # Fill in the External Links values
    source_url_col = df.columns.get_loc('Source URL') + 1
    for row_idx, row in enumerate(df.itertuples(), 1):
        if hasattr(row, 'Source_URL'):
            source_url = getattr(row, 'Source_URL')
            if source_url in source_url_to_ext_links:
                ext_links_value = source_url_to_ext_links[source_url]
                worksheet.cell(
                    row=start_row + row_idx,
                    column=ext_links_col,
                    value=ext_links_value
                )

    return True