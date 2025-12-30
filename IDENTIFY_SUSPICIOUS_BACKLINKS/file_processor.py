# file_processor.py
import pandas as pd
import traceback
from openpyxl.styles import Font, Alignment, PatternFill
from core.utils import standardize_semrush_columns, analyze_last_seen_dates, sanitize_sheet_name, extract_domain_from_url
from openpyxl.utils import get_column_letter
from content_analysis.detection import analyze_content
from content_analysis.reporting import add_dodgy_domain_sheet, create_valid_sheet_name

# Import from new module structure - with corrected import paths
try:
    from excel_utils import add_df_to_worksheet, add_navigation_links
    # Updated import from navigating instead of navigation
    from excel_utils.navigating import fix_invalid_hyperlinks
except ImportError:
    # Fallback to old imports
    from excel_utils import add_df_to_worksheet, add_navigation_links
    # Let this fail if needed, as we're fixing the structure


def read_backlink_file(file_path):
    """Read a backlink file (CSV or Excel)"""
    import pandas as pd
    try:
        if file_path.lower().endswith('.csv'):
            df = pd.read_csv(file_path)
        else:
            df = pd.read_excel(file_path)
        print(f"Loaded data with {len(df)} rows.")
        return df
    except Exception as e:
        print(f"Error reading file {file_path}: {e}")
        return None


# Import configuration
try:
    from config import FEATURES
except ImportError:
    # Default features if config not available
    FEATURES = {
        'SORT_DOMAINS_BY_SIZE': True,
        'GROUP_SIMILAR_BACKLINKS': True,
        'ADD_HYPERLINKS': True,
        'USE_DOMAIN_SPACING': True,
    }

# Import configuration if available
try:
    from config import QUALITY_BACKLINK_SETTINGS, SPAMMY_BACKLINK_SETTINGS, FEATURES
except ImportError:
    # Default settings if config not available
    QUALITY_BACKLINK_SETTINGS = {
        'MIN_AUTHORITY_SCORE': 5,
        'MAX_EXTERNAL_LINKS': 200,
        'REQUIRE_DOFOLLOW': True
    }
    SPAMMY_BACKLINK_SETTINGS = {
        'MAX_AUTHORITY_SCORE': 5,
        'MIN_EXTERNAL_LINKS': 200,
        'MIN_AUTHORITY_SCORE': 0
    }
    FEATURES = {
        'GROUP_SIMILAR_BACKLINKS': True,
        'ENABLE_TRANSLATION': False,  # Translation disabled
        'REORDER_COLUMNS': True,
        'SORT_BY_DOMAIN_COUNT': True
    }


def add_referring_domain_column(df):
    """Add a column with the referring domain extracted from the source URL"""
    if 'Source url' in df.columns:
        df['Referring Domain'] = df['Source url'].apply(extract_domain_from_url)
    return df


def add_domain_backlinks_count(df):
    """Add a column showing the number of backlinks per domain"""
    if 'Referring Domain' in df.columns:
        # Count backlinks per domain
        domain_counts = df['Referring Domain'].value_counts().to_dict()

        # Track seen domains to only show count for first occurrence
        seen_domains = set()
        domain_count_col = []

        for domain in df['Referring Domain']:
            if domain and domain not in seen_domains:
                seen_domains.add(domain)
                domain_count_col.append(domain_counts.get(domain, 0))
            else:
                domain_count_col.append('')

        # Insert new column before Referring Domain
        domain_col_idx = df.columns.get_loc('Referring Domain')
        df.insert(domain_col_idx, 'Domain Backlinks', domain_count_col)

    return df


def create_summary_statistics(df):
    """Create additional summary statistics for Semrush data"""
    stats = {}

    # Count total links
    stats['Total Links'] = len(df)

    # Count dofollow vs nofollow
    if 'Nofollow' in df.columns:
        dofollow_count = len(df[~df['Nofollow']])
        stats['Dofollow Links'] = dofollow_count
        stats['Nofollow Links'] = len(df) - dofollow_count

    # Count lost links
    if 'Lost link' in df.columns:
        if df['Lost link'].dtype == bool:
            # If it's a boolean column
            lost_links = df['Lost link']
        else:
            # If it's a string or other type
            lost_links = df['Lost link'].notna() & (df['Lost link'] != '') & (df['Lost link'] != 'False')

        stats['Lost Links'] = lost_links.sum()
        stats['Active Links'] = len(df) - lost_links.sum()

    # Average Page ascore
    if 'Page ascore' in df.columns:
        stats['Avg Page ascore'] = df['Page ascore'].mean()

    return stats


def group_similar_backlinks(df, enable_grouping=True):
    """Group backlinks that have the same source title, URL, and anchor text"""
    if not enable_grouping:
        return df

    if 'Source title' in df.columns and 'Source url' in df.columns and 'Anchor' in df.columns:
        try:
            # Group by the key columns
            grouped = df.groupby(['Source title', 'Source url', 'Anchor']).agg({
                'Page ascore': 'mean',  # Average score for grouped links
                'External links': 'mean',  # Average external links
                'First seen': 'min',  # Earliest first seen date
                'Last seen': 'max',  # Latest last seen date
                'Target url': 'first',  # Keep one target URL
                'Nofollow': 'any',  # If any are nofollow, mark as nofollow
                'Sitewide': 'any',  # If any are sitewide, mark as sitewide
                'Lost link': 'any',  # If any are lost, mark as lost
                'Referring Domain': 'first'  # Keep the domain
            }).reset_index()

            # Round numeric columns to keep them clean
            for col in ['Page ascore', 'External links']:
                if col in grouped.columns:
                    grouped[col] = grouped[col].round(1)

            # Add count column to show frequency
            counts = df.groupby(['Source title', 'Source url', 'Anchor']).size().reset_index(name='Frequency')
            grouped = grouped.merge(counts, on=['Source title', 'Source url', 'Anchor'])

            # Sort by frequency in descending order
            grouped = grouped.sort_values('Frequency', ascending=False)

            print(f"Grouped backlinks from {len(df)} to {len(grouped)} rows")
            return grouped
        except Exception as e:
            print(f"Error grouping backlinks: {e}")
            return df

    return df


def get_domain_everything_df(file_path):
    """Extract just the Everything dataframe from a backlink file without creating sheets"""
    df = read_backlink_file(file_path)
    if df is None:
        return None, None

    # Apply the same preprocessing as in process_backlink_file
    from utils import standardize_semrush_columns, remove_translation_references
    df = standardize_semrush_columns(df)
    # Remove any translation-related columns
    df = remove_translation_references(df)
    df = add_referring_domain_column(df)
    df = add_domain_backlinks_count(df)

    # Extract domain name from the file path
    from core.utils import extract_domain_name
    domain = extract_domain_name(file_path)  # Get domain name from file path

    # Group similar backlinks if enabled
    if FEATURES.get('GROUP_SIMILAR_BACKLINKS', True):
        df = group_similar_backlinks(df, enable_grouping=True)

    # Return the DataFrame
    return df, None


def standardize_column_names(df):
    """Standardize column names to ensure consistency"""
    if df is None or df.empty:
        return df
        
    # Create a mapping of possible column names to our standard names
    column_map = {
        'Source URL': 'Source url',
        'Source Url': 'Source url',
        'Source': 'Source url',
        'Page AS': 'Page ascore',
        'Page Authority Score': 'Page ascore',
        'Authority Score': 'Page ascore',
        'AS': 'Page ascore',
        'Target': 'Target url',
        'Target Url': 'Target url',
        'Target URL': 'Target url',
        'Title': 'Source title',
        'Source Title': 'Source title',
        'FirstSeen': 'First seen',
        'First Seen': 'First seen',
        'LastSeen': 'Last seen',
        'Last Seen': 'Last seen',
        'ExtBacklinks': 'External links',
        'External Links': 'External links',
        'External backlinks': 'External links',
        'NoFollow': 'Nofollow',
        'No Follow': 'Nofollow',
        'SiteWide': 'Sitewide',
        'Site Wide': 'Sitewide',
        'Lost Link': 'Lost link',
        'LostLink': 'Lost link'
    }
    
    # Rename columns that match our mapping
    df = df.rename(columns=column_map)
    return df


def process_backlink_file(workbook, file_path, domain_name, sheet_number=None):
    """
    Process a single backlink file and add it to the workbook.
    """
    # Create valid sheet names
    everything_sheet_name = create_valid_sheet_name(domain_name, "_Everything")
    dodgy_sheet_name = create_valid_sheet_name(domain_name, "_dodgy")
    
    import logging
    import traceback
    from pandas import DataFrame
    
    # Set up logging
    logging.basicConfig(level=logging.INFO)
    logger = logging.getLogger(__name__)
    
    try:
        # Validate domain
        if not domain_name:
            logger.warning(f"No domain name provided for {file_path}")
            return False, None, None
            
        # Read the file
        logger.info(f"Reading file: {file_path}")
        df = read_backlink_file(file_path)
        
        # Validate dataframe
        if df is None:
            logger.error(f"Failed to read file: {file_path}")
            return False, None, None
            
        if not isinstance(df, DataFrame):
            logger.error(f"Invalid data type returned from read_backlink_file: {type(df)}")
            return False, None, None
            
        if df.empty:
            logger.warning(f"Empty dataframe in file: {file_path}")
            return False, None, None
            
        # Standardize column names
        logger.info("Standardizing column names")
        df = standardize_column_names(df)
        
        # Validate required columns
        required_columns = ['Source url', 'Target url', 'Page ascore', 'External links']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            logger.error(f"Missing required columns in {file_path}: {missing_columns}")
            return False, None, None
            
        # Add referring domain column
        logger.info("Adding referring domain column")
        df = add_referring_domain_column(df)
        
        # Validate domain data
        if 'Referring Domain' not in df.columns:
            logger.error(f"No domain data found in {file_path}")
            return False, None, None
            
        # Add domain backlinks count
        logger.info("Adding domain backlinks count")
        df = add_domain_backlinks_count(df)
        
        # Analyze content for suspicious keywords
        logger.info("Analyzing content for suspicious keywords")
        domain_results = analyze_content(df, domain_name=domain_name)
        
        # If dodgy keywords found, create a domain-specific dodgy sheet
        if not domain_results.empty:
            add_dodgy_domain_sheet(workbook, domain_name, domain_results, df, sheet_number)
            logger.info(f"Created dodgy sheet for domain: {domain_name}")
        
        logger.info(f"Successfully processed file: {file_path}")
        return True, None, None
        
    except Exception as e:
        logger.error(f"Error processing file {file_path}: {str(e)}")
        logger.error("Traceback:", exc_info=True)
        return False, None, None


def add_df_to_worksheet(worksheet, df):
    """Add a DataFrame to a worksheet with proper formatting"""
    if df is None or df.empty:
        return
        
    # Convert DataFrame to list of lists
    data = [df.columns.tolist()] + df.values.tolist()
    
    # Write data to worksheet
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
            
            # Format header row
            if row_idx == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
    # Auto-adjust column widths
    for col in range(1, len(df.columns) + 1):
        max_length = 0
        column = get_column_letter(col)
        for cell in worksheet[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap width at 50 characters
        worksheet.column_dimensions[column].width = adjusted_width
        
    # Freeze the header row
    worksheet.freeze_panes = 'A2'


def create_time_summary(df):
    """Create a summary of time-based metrics for backlinks"""
    time_summary = {}

    # Check if we have the necessary columns
    if 'First seen' in df.columns and 'Last seen' in df.columns:
        try:
            # Convert to datetime if not already
            df['First seen'] = pd.to_datetime(df['First seen'])
            df['Last seen'] = pd.to_datetime(df['Last seen'])

            # Get earliest and latest dates
            time_summary['Earliest Link'] = df['First seen'].min()
            time_summary['Latest Link'] = df['First seen'].max()
            time_summary['Most Recent Update'] = df['Last seen'].max()

            # Count links by age
            now = pd.Timestamp.now()
            df['Age'] = (now - df['First seen']).dt.days

            # Define age buckets (in days)
            age_buckets = {
                '0-30 days': 30,
                '31-90 days': 90,
                '91-180 days': 180,
                '181-365 days': 365,
                '1-2 years': 730,
                'Over 2 years': float('inf')
            }

            # Count links in each age bucket
            prev_limit = 0
            for label, limit in age_buckets.items():
                count = len(df[(df['Age'] > prev_limit) & (df['Age'] <= limit)])
                time_summary[f'Links {label}'] = count
                prev_limit = limit

        except Exception as e:
            print(f"Error creating time summary: {e}")

    return time_summary